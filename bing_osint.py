import requests
from bs4 import BeautifulSoup
from datetime import datetime
import urllib.parse
import time
from requests.packages.urllib3.exceptions import InsecureRequestWarning
import pandas as pd
import openpyxl
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter

requests.packages.urllib3.disable_warnings(InsecureRequestWarning)         

def extract_bing_results(html):
    soup = BeautifulSoup(html, 'html.parser')
    results = []

    for li in soup.select("li.b_algo"):
        a_tag = li.find("h2").find("a")
        desc_tag = li.select_one("div.b_caption > p") 
        if a_tag:
            title = a_tag.get_text(strip=True)
            href = a_tag.get("href")
            content = desc_tag.get_text(strip=True) if desc_tag else ""
            results.append((title, href, content))

    return results

def bing_search(query):
    headers = {
    "Host": "www.bing.com",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/114.0.0.0 Safari/537.36",
    "Cookie": "MUIDB=33EEB804607960B83BA9ADE9610E6120; SRCHHPGUSR=SRCHLANG=ko&IG=BFE2C78C421D4C57AD47F0752E1CFCE2&PV=10.0.0&DM=0&BRW=M&BRH=S&CW=1298&CH=551&SCW=1402&SCH=1503&DPR=1.0&UTC=540&HV=1751679538&HVE=CfDJ8Inh5QCoSQBNls38F2rbEpTuvls9C0V30J0pv1GBjDav1a95uJGhoVyTrfiz8Udp4NgQR0IpKSHaAV6z5CmkDjPy8V-urizCgWhiBd95XOzDieG792t4McwyYgHbh-MPL7tTideou7WzflQhd2ckljwAg4c0JdtyiRczvtRR1UO-vcMQrGAahnXU8GlUin030A&PRVCW=1298&PRVCH=551&EXLTT=7"
}
    results = []
    offset = 0

    while True:
        q = urllib.parse.quote(query)
        url = f"https://www.bing.com/search?q={q}&first={offset}&setLang=ko&cc=KR"
        res = requests.get(url, headers=headers, verify=False)
        soup = BeautifulSoup(res.text, 'html.parser')

        extracted = extract_bing_results(res.text)
        for title, url, content in extracted:
            print(f"[{title}]\n{url}\n{content}")
        if "다음에 대한 결과가 없음" in res.text:
            break

        offset += 10
        time.sleep(1)  # polite delay to avoid being blocked
        results.append(extracted)
    return results


if __name__ == "__main__":
    timestamp = datetime.now().strftime("%m%d_%H%M")
    filename =  f"BingSearchResult_{timestamp}"
    search = input("search(ex: site:https://google.com...)>")
    search = search.replace("+", " ")
    result = bing_search(search)
    flat_result = [item for sublist in result for item in sublist]

    df = pd.DataFrame(flat_result, columns=["제목", "링크", "내용"])
    df.index += 1  
    df.to_csv(filename+".csv", index_label="No", encoding="utf-8-sig")
    df.to_excel(filename+".xlsx", index_label="No")
    wb = openpyxl.load_workbook(filename+".xlsx")
    ws = wb.active

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border

    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_length + 2  

    wb.save(filename+".xlsx")